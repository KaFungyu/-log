import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 文件夹路径（请修改为您实际的路径）
folder_path = r'D:\Desktop\666\USB2.0 host HS'

# 创建Excel工作簿和表格
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Eye Diagram"

# 定义图片在Excel中的大小，单位为厘米
img_height_cm = 2.2  # cm
img_width_cm = 4  # cm

# 将厘米转换为像素，1 cm = 37.795275591 像素
img_height_px = img_height_cm * 37.795275591
img_width_px = img_width_cm * 37.795275591

# 定义标题行和路径、名称（可以根据需要修改）
column_headers = ["FF1#4", "FF1#5", "FF1#6", "TT#8", "TT#9", "TT#10", "SS2#6", "SS2#7", "SS2#9"]
row_headers = ["HTHV", "HTLV", "LTHV", "LTLV", "NTHV", "NTLV"]

# 写入标题行
first_row = ["Eye Diagram"] + column_headers
ws.append(first_row)

# 写入路径行
for row_header in row_headers:
    ws.append([row_header])

# 计算表格的实际大小
total_rows = len(row_headers) + 1  # 加1是因为有标题行
total_cols = len(column_headers) + 1  # 加1是因为有第一列

# 列宽设置，A列宽度设置为10.5字符，其余列的宽度设置为19字符
ws.column_dimensions['A'].width = 10.5
for col_idx in range(2, total_cols + 1):  # 从B列开始（索引2）
    col_letter = chr(64 + col_idx)  # 将数字索引转换为字母（A=65, B=66...）
    ws.column_dimensions[col_letter].width = 19

# 行高设置，根据实际行数动态设置行高
for row_idx in range(2, total_rows + 1):  # 从第2行开始
    ws.row_dimensions[row_idx].height = 65  # 设置为磅

# 设置字体和对齐方式
font = Font(name='微软雅黑', size=10)
alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

# 设置边框样式
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# 设置所有单元格的字体、对齐方式和边框
for row in range(1, total_rows + 1):
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = alignment
        cell.border = border

# 设置背景填充颜色（标准橙色）
yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")

# 为标题行（除A1外）设置背景色
for col in range(2, total_cols + 1):
    ws.cell(row=1, column=col).fill = yellow_fill

# 为第一列（除A1外）设置背景色
for row in range(2, total_rows + 1):
    ws.cell(row=row, column=1).fill = yellow_fill

# 统计填充的图片数量
image_count = 0

# 遍历文件夹中的所有图片
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file.lower().endswith('2.png'):
            # 解析路径和文件名
            file_path = os.path.join(root, file)
            folder_name = os.path.basename(root)  # 获取文件夹名，如 FF1#4
            path_name = os.path.basename(os.path.dirname(root))  # 获取路径名，如 HTHV
            
            # 打印调试信息
            print(f"Processing image: {file_path}")
            print(f"Path name: {path_name}, Folder name: {folder_name}")

            # 检查路径名和文件夹名是否在我们的标题中
            if path_name not in row_headers or folder_name not in column_headers:
                print(f"跳过不匹配的图片: 路径名 '{path_name}' 或文件夹名 '{folder_name}' 不在定义的标题中")
                continue

            try:
                img_obj = Image(file_path)
                img_obj.width = img_width_px
                img_obj.height = img_height_px

                # 直接通过索引确定单元格位置，提高性能
                row_idx = row_headers.index(path_name) + 2  # +2 因为索引从0开始，且第一行是标题
                col_idx = column_headers.index(folder_name) + 2  # +2 因为索引从0开始，且第一列是路径名
                
                cell = f'{chr(64 + col_idx)}{row_idx}'
                ws.add_image(img_obj, cell)

                # 更新图片计数
                image_count += 1
                print(f"Added image to Excel at {cell}: {file_path}")

            except Exception as e:
                print(f"Error adding image to Excel: {e}")

# 打印填充的图片总数
print(f"Total images added: {image_count}")

# 保存Excel文件
excel_output_path = r'D:\Desktop\666\usb2.0 HS host_eye.xlsx'
wb.save(excel_output_path)
print(f"处理完成，文件已保存至: {excel_output_path}")

# 自动打开生成的Excel文件
try:
    os.startfile(excel_output_path)
    print(f"已自动打开文件: {excel_output_path}")
except Exception as e:
    print(f"无法自动打开文件: {e}")
