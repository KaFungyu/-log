import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 基础文件夹路径（请修改为实际路径）
base_folder_path = r'D:\Desktop\666\eth'

# 创建Excel工作簿和表格
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "100M eth eye"

# 定义标题行和路径、名称（可以根据需要修改）
column_headers = ["FF1#4", "FF1#5", "FF1#6", "TT#8", "TT#9", "TT#10", "SS2#6", "SS2#7", "SS2#9"]
row_headers = ["HTHV", "HTLV", "LTHV", "LTLV", "NTHV", "NTLV"]

# 写入标题行
first_row = ["Eye Top"] + column_headers
ws.append(first_row)

# 写入路径行
for row_header in row_headers:
    ws.append([row_header])

# 定义图片在Excel中的大小，单位为厘米
img_height_cm = 2.2  # cm
img_width_cm = 4  # cm

# 将厘米转换为像素，1 cm = 37.795275591 像素
img_height_px = img_height_cm * 37.795275591
img_width_px = img_width_cm * 37.795275591

# 设置Excel表格的列宽和行高
ws.column_dimensions['A'].width = 10.5  # 第一列固定宽度

# 根据实际列数动态设置列宽
for col_idx in range(2, len(column_headers) + 2):  # 从B列开始（索引2）
    col_letter = chr(64 + col_idx)  # 将数字索引转换为字母（A=65, B=66...）
    ws.column_dimensions[col_letter].width = 19

# 行高将在所有操作完成后统一设置

# 设置字体和对齐方式
font = Font(name='微软雅黑', size=10)
alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

# 设置边框样式
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# 计算表格的实际大小
total_rows = len(row_headers) + 1  # 加1是因为有标题行
total_cols = len(column_headers) + 1  # 加1是因为有第一列

# 设置所有单元格的字体、对齐方式和边框（包括上半部分和下半部分）
for row in range(1, 2 * total_rows + 1):
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = alignment
        cell.border = border

# 设置背景颜色（橙色）
yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
# 为上半部分标题行（除A1外）设置背景色
for col in range(2, total_cols + 1):
    ws.cell(row=1, column=col).fill = yellow_fill
# 为上半部分第一列（除A1外）设置背景色
for row in range(2, total_rows + 1):
    ws.cell(row=row, column=1).fill = yellow_fill
    
# 为下半部分标题行（除A1外）设置背景色
for col in range(2, total_cols + 1):
    ws.cell(row=total_rows+1, column=col).fill = yellow_fill
# 为下半部分第一列（除A1外）设置背景色
for row in range(total_rows+2, 2*total_rows + 1):
    ws.cell(row=row, column=1).fill = yellow_fill

# 统计填充的图片数量
image_count = 0

# 遍历所有路径和文件夹名进行图片填充
for row_idx, path_value in enumerate(row_headers, 2):  # 从第2行开始，对应路径名
    for col_idx, folder_value in enumerate(column_headers, 2):  # 从第2列开始，对应文件夹名
        folder_path_combined = os.path.join(base_folder_path, path_value, folder_value)

        # 检查路径是否存在
        if os.path.exists(folder_path_combined):
            print(f"Processing folder: {folder_path_combined}")

            # 查找该路径和文件夹中的所有以 '28' 开头的 .png 文件
            dir_files = [f for f in os.listdir(folder_path_combined) if f.startswith('28') and f.endswith('.png')]

            if dir_files:
                # 根据生成时间对文件进行排序
                file_paths_with_times = [(os.path.join(folder_path_combined, f), os.path.getctime(os.path.join(folder_path_combined, f))) for f in dir_files]
                file_paths_with_times.sort(key=lambda x: x[1])  # 按照生成时间升序排序

                # 处理较早的图片，填充到上半部分
                if len(file_paths_with_times) > 0:
                    early_image_path, _ = file_paths_with_times[0]
                    img_obj = Image(early_image_path)
                    img_obj.width = img_width_px
                    img_obj.height = img_height_px
                    ws.add_image(img_obj, f"{chr(64 + col_idx)}{row_idx}")

                    image_count += 1
                    print(f"Added early image to {chr(64 + col_idx)}{row_idx}: {early_image_path}")

                # 处理较晚的图片，填充到下半部分
                if len(file_paths_with_times) > 1:
                    late_image_path, _ = file_paths_with_times[-1]
                    img_obj = Image(late_image_path)
                    img_obj.width = img_width_px
                    img_obj.height = img_height_px
                    ws.add_image(img_obj, f"{chr(64 + col_idx)}{row_idx + total_rows}")

                    image_count += 1
                    print(f"Added late image to {chr(64 + col_idx)}{row_idx + total_rows}: {late_image_path}")

        else:
            print(f"Folder not found: {folder_path_combined}")

# 复制上半部分内容到下半部分
for row in range(1, total_rows + 1):
    for col in range(1, total_cols + 1):
        source_cell = ws.cell(row=row, column=col)
        target_cell = ws.cell(row=row + total_rows, column=col)

        # 复制内容
        if row == 1 and col == 1:
            target_cell.value = "Eye Bottom"  # 修改下半部分第一个单元格的值为"Eye Bottom"
        else:
            target_cell.value = source_cell.value

        # 复制字体、对齐方式、边框和填充
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )

# 确保所有行高设置正确应用（放在所有操作的最后）
# 上半部分行高设置
for row_idx in range(2, total_rows + 1):
    ws.row_dimensions[row_idx].height = 65

# 下半部分行高设置
for row_idx in range(total_rows + 1, 2 * total_rows + 1):
    ws.row_dimensions[row_idx].height = 65

# 打印填充的图片总数
print(f"Total images added: {image_count}")

# 保存Excel文件
excel_output_path = r'D:\Desktop\666\100M_eth_eye.xlsx'
wb.save(excel_output_path)
print(f"处理完成，文件已保存至: {excel_output_path}")

# 自动打开生成的Excel文件
try:
    os.startfile(excel_output_path)
    print(f"已自动打开文件: {excel_output_path}")
except Exception as e:
    print(f"无法自动打开文件: {e}")